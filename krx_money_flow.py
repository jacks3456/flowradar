#!/usr/bin/env python3
"""Fetch KRX investor money flow and generate Excel and HTML reports."""

from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

os.environ.setdefault("MPLCONFIGDIR", "/private/tmp/krx_matplotlib")

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill


KRX_JSON_URL = "https://data.krx.co.kr/comm/bldAttendant/getJsonData.cmd"
REFERER = "https://data.krx.co.kr/contents/MDC/MDI/outerLoader/index.cmd"
LOGIN_PAGE = "https://data.krx.co.kr/contents/MDC/COMS/client/MDCCOMS001.cmd"
LOGIN_JSP = "https://data.krx.co.kr/contents/MDC/COMS/client/view/login.jsp?site=mdc"
LOGIN_URL = "https://data.krx.co.kr/contents/MDC/COMS/client/MDCCOMS001D1.cmd"

BLD_STOCK_SEARCH = "dbms/comm/finder/finder_stkisu"
BLD_MARKET_DAILY = "dbms/MDC/STAT/standard/MDCSTAT02202"
BLD_STOCK_DAILY = "dbms/MDC/STAT/standard/MDCSTAT02302"

INVESTOR_COLUMNS = {
    "TRDVAL1": "机构",
    "TRDVAL2": "其他法人",
    "TRDVAL3": "散户",
    "TRDVAL4": "外资",
    "TRDVAL_TOT": "合计",
}

DEFAULT_TICKERS = {
    "005930": "三星电子",
    "000660": "SK海力士",
}


@dataclass(frozen=True)
class StockInfo:
    ticker: str
    name: str
    isin: str
    market: str


class KrxClient:
    def __init__(
        self,
        pause_seconds: float = 0.2,
        login_id: str | None = None,
        login_password: str | None = None,
        cookie_header: str | None = None,
    ) -> None:
        self.pause_seconds = pause_seconds
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "Mozilla/5.0",
                "Referer": REFERER,
                "X-Requested-With": "XMLHttpRequest",
            }
        )
        self.login_id = login_id or os.getenv("KRX_ID")
        self.login_password = login_password or os.getenv("KRX_PW")
        self.cookie_header = cookie_header or os.getenv("KRX_COOKIE")
        if self.cookie_header:
            self.session.headers.update({"Cookie": self.cookie_header})
        elif self.login_id and self.login_password:
            self.login()

    def login(self) -> None:
        self.session.get(LOGIN_PAGE, timeout=30)
        self.session.get(LOGIN_JSP, headers={"Referer": LOGIN_PAGE}, timeout=30)
        payload = {
            "mbrNm": "",
            "telNo": "",
            "di": "",
            "certType": "",
            "mbrId": self.login_id,
            "pw": self.login_password,
        }
        response = self.session.post(
            LOGIN_URL,
            data=payload,
            headers={"Referer": LOGIN_PAGE},
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()
        error_code = data.get("_error_code", "")
        if error_code == "CD011":
            payload["skipDup"] = "Y"
            response = self.session.post(
                LOGIN_URL,
                data=payload,
                headers={"Referer": LOGIN_PAGE},
                timeout=30,
            )
            response.raise_for_status()
            data = response.json()
            error_code = data.get("_error_code", "")

        if error_code != "CD001":
            message = data.get("_error_message", "KRX login failed")
            raise RuntimeError(f"KRX 登录失败: {error_code} {message}")

    def post_json(self, bld: str, **params: object) -> dict:
        payload = {"bld": bld, "locale": "ko_KR"}
        payload.update(params)
        response = self.session.post(KRX_JSON_URL, data=payload, timeout=30)
        if response.status_code == 400 and "LOGOUT" in response.text:
            raise RuntimeError(
                "KRX 返回 LOGOUT。Google 登录用户请从浏览器复制 KRX 已登录 cookie，"
                "通过 KRX_COOKIE 环境变量或 --krx-cookie 传入；普通账号可设置 "
                "KRX_ID/KRX_PW 或使用 --krx-id/--krx-password。"
            )
        response.raise_for_status()
        time.sleep(self.pause_seconds)
        return response.json()

    def listed_stocks(self, market: str = "ALL") -> list[StockInfo]:
        data = self.post_json(
            BLD_STOCK_SEARCH,
            mktsel=market,
            searchText="",
            typeNo=0,
        )
        rows = data.get("block1", [])
        return [
            StockInfo(
                ticker=row["short_code"],
                name=row["codeName"],
                isin=row["full_code"],
                market=row.get("marketEngName") or row.get("marketName", ""),
            )
            for row in rows
        ]

    def market_daily_flow(
        self,
        start: str,
        end: str,
        market: str = "KOSPI",
        include_etf: bool = False,
        include_etn: bool = False,
        include_elw: bool = False,
    ) -> pd.DataFrame:
        market_id = market_to_krx_id(market)
        data = self.post_json(
            BLD_MARKET_DAILY,
            strtDd=start,
            endDd=end,
            mktId=market_id,
            etf="EF" if include_etf else "",
            etn="EN" if include_etn else "",
            elw="EW" if include_elw else "",
            inqTpCd=2,
            trdVolVal=2,
            askBid=3,
        )
        return normalize_flow(data.get("output", []), value_unit="eok_krw")

    def stock_daily_flow(
        self,
        start: str,
        end: str,
        isin: str,
        metric: str,
    ) -> pd.DataFrame:
        data = self.post_json(
            BLD_STOCK_DAILY,
            strtDd=start,
            endDd=end,
            isuCd=isin,
            inqTpCd=2,
            trdVolVal=1 if metric == "shares" else 2,
            askBid=3,
        )
        unit = "shares" if metric == "shares" else "krw"
        return normalize_flow(data.get("output", []), value_unit=unit)


def parse_args() -> argparse.Namespace:
    today = dt.date.today()
    default_start = today - dt.timedelta(days=90)
    parser = argparse.ArgumentParser(
        description="KRX 外资/机构/散户资金流 Excel 报表生成器"
    )
    parser.add_argument("--start", default=format_date(default_start), help="开始日期 YYYYMMDD")
    parser.add_argument("--end", default=format_date(today), help="结束日期 YYYYMMDD")
    parser.add_argument("--market", default="KOSPI", choices=["KOSPI", "KOSDAQ", "KONEX", "ALL"])
    parser.add_argument(
        "--tickers",
        default="005930,000660",
        help="逗号分隔股票代码，例如 005930,000660",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出 Excel 路径，默认 reports/krx_money_flow_开始_结束.xlsx",
    )
    parser.add_argument(
        "--html-output",
        default=None,
        help="输出网页路径，默认与 Excel 同名 .html",
    )
    parser.add_argument(
        "--site-dir",
        default="site",
        help="公开网站目录，默认 site；脚本会写入 site/data/latest.json",
    )
    parser.add_argument(
        "--no-site",
        action="store_true",
        help="不生成公开网站数据源",
    )
    parser.add_argument("--include-etf", action="store_true", help="市场资金流包含 ETF")
    parser.add_argument("--include-etn", action="store_true", help="市场资金流包含 ETN")
    parser.add_argument("--include-elw", action="store_true", help="市场资金流包含 ELW")
    parser.add_argument("--krx-id", default=None, help="KRX 登录账号，也可用 KRX_ID 环境变量")
    parser.add_argument(
        "--krx-password",
        default=None,
        help="KRX 登录密码，也可用 KRX_PW 环境变量",
    )
    parser.add_argument(
        "--krx-cookie",
        default=None,
        help="已登录浏览器里的 KRX Cookie 字符串，也可用 KRX_COOKIE 环境变量",
    )
    return parser.parse_args()


def format_date(value: dt.date) -> str:
    return value.strftime("%Y%m%d")


def market_to_krx_id(market: str) -> str:
    return {
        "KOSPI": "STK",
        "KOSDAQ": "KSQ",
        "KONEX": "KNX",
        "ALL": "ALL",
    }[market.upper()]


def normalize_number(value: object) -> int:
    text = str(value).strip()
    if text in {"", "-"}:
        return 0
    text = re.sub(r"[^\d.-]", "", text)
    if text in {"", "-"}:
        return 0
    return int(float(text))


def normalize_flow(rows: list[dict], value_unit: str) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=["日期", "机构", "其他法人", "散户", "外资", "合计"])

    df = pd.DataFrame(rows).rename(columns=INVESTOR_COLUMNS)
    keep = ["TRD_DD", "机构", "其他法人", "散户", "外资", "合计"]
    df = df[[col for col in keep if col in df.columns]].rename(columns={"TRD_DD": "日期"})
    df["日期"] = pd.to_datetime(df["日期"], format="%Y/%m/%d")

    value_columns = [col for col in df.columns if col != "日期"]
    for col in value_columns:
        df[col] = df[col].map(normalize_number)

    df["合计"] = df[["机构", "散户", "外资"]].sum(axis=1)
    value_columns = [col for col in df.columns if col != "日期"]

    if value_unit == "eok_krw":
        for col in value_columns:
            df[col] = df[col] / 100_000_000
    elif value_unit not in {"krw", "shares"}:
        raise ValueError(f"Unknown value unit: {value_unit}")

    return df.sort_values("日期").reset_index(drop=True)


def current_streak(values: Iterable[float]) -> int:
    cleaned = [value for value in values if value != 0]
    if not cleaned:
        return 0
    sign = 1 if cleaned[-1] > 0 else -1
    streak = 0
    for value in reversed(cleaned):
        if (value > 0 and sign > 0) or (value < 0 and sign < 0):
            streak += 1
        else:
            break
    return streak * sign


def streak_label(streak: int) -> str:
    if streak > 0:
        return f"连续净买入 {streak} 天"
    if streak < 0:
        return f"连续净卖出 {abs(streak)} 天"
    return "无连续方向"


def build_summary(
    market_flow: pd.DataFrame,
    stock_flows: dict[str, dict[str, pd.DataFrame]],
    market: str = "KOSPI",
) -> pd.DataFrame:
    rows = []
    market_object = f"{market.upper()}市场"
    for investor in ["外资", "机构", "散户"]:
        streak = current_streak(market_flow[investor])
        rows.append(
            {
                "对象": market_object,
                "指标": f"{investor}净买卖金额",
                "最近值": market_flow[investor].iloc[-1] if not market_flow.empty else 0,
                "单位": "亿韩元",
                "连续状态": streak_label(streak),
            }
        )

    for ticker, frames in stock_flows.items():
        for investor in ["外资", "机构", "散户"]:
            shares = frames["shares"]
            value = frames["value"]
            share_streak = current_streak(shares[investor])
            value_streak = current_streak(value[investor])
            rows.append(
                {
                    "对象": ticker,
                    "指标": f"{investor}净买卖股数",
                    "最近值": shares[investor].iloc[-1] if not shares.empty else 0,
                    "单位": "股",
                    "连续状态": streak_label(share_streak),
                }
            )
            rows.append(
                {
                    "对象": ticker,
                    "指标": f"{investor}净买卖金额",
                    "最近值": value[investor].iloc[-1] if not value.empty else 0,
                    "单位": "韩元",
                    "连续状态": streak_label(value_streak),
                }
            )
    return pd.DataFrame(rows)


def plot_market_flow(market_flow: pd.DataFrame, chart_path: Path) -> None:
    import matplotlib.pyplot as plt

    chart_labels = {"外资": "Foreign", "机构": "Institution", "散户": "Retail"}
    plt.rcParams["axes.unicode_minus"] = False
    plt.figure(figsize=(11, 5.5))
    for investor in ["外资", "机构", "散户"]:
        plt.plot(
            market_flow["日期"],
            market_flow[investor].cumsum(),
            label=chart_labels[investor],
            linewidth=2,
        )
    plt.axhline(0, color="#777777", linewidth=0.8)
    plt.title("KOSPI Investor Net Buying Cumulative Flow")
    plt.ylabel("100M KRW")
    plt.grid(True, alpha=0.25)
    plt.legend()
    plt.tight_layout()
    chart_path.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(chart_path, dpi=180)
    plt.close()


def write_report(
    output_path: Path,
    market_flow: pd.DataFrame,
    stock_flows: dict[str, dict[str, pd.DataFrame]],
    summary: pd.DataFrame,
    chart_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        market_flow.to_excel(writer, sheet_name="kospi_market_eok_krw", index=False)
        for ticker, frames in stock_flows.items():
            frames["shares"].to_excel(writer, sheet_name=f"{ticker}_shares", index=False)
            frames["value"].to_excel(writer, sheet_name=f"{ticker}_krw", index=False)

    workbook = load_workbook(output_path)
    for sheet in workbook.worksheets:
        format_sheet(sheet)

    sheet = workbook.create_sheet("charts")
    sheet["A1"] = "KOSPI 外资/机构/散户累计净买卖曲线"
    sheet.add_image(ExcelImage(chart_path), "A3")
    workbook.save(output_path)


def write_html_dashboard(
    html_path: Path,
    market_flow: pd.DataFrame,
    stock_flows: dict[str, dict[str, pd.DataFrame]],
    summary: pd.DataFrame,
    *,
    start: str,
    end: str,
    market: str,
) -> None:
    html_path.parent.mkdir(parents=True, exist_ok=True)
    payload = build_dashboard_payload(
        market_flow,
        stock_flows,
        summary,
        start=start,
        end=end,
        market=market,
    )
    data_json = json.dumps(payload, ensure_ascii=False)
    html_path.write_text(build_html(data_json, payload["meta"]), encoding="utf-8")


def write_site_data(site_dir: Path, payload: dict) -> Path:
    data_path = site_dir / "data" / "latest.json"
    data_path.parent.mkdir(parents=True, exist_ok=True)
    data_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return data_path


def build_dashboard_payload(
    market_flow: pd.DataFrame,
    stock_flows: dict[str, dict[str, pd.DataFrame]],
    summary: pd.DataFrame,
    *,
    start: str,
    end: str,
    market: str,
) -> dict:
    return {
        "meta": {
            "start": start,
            "end": end,
            "market": market,
            "generatedAt": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "latestDate": latest_date_label(market_flow),
        },
        "summary": frame_to_records(summary),
        "marketFlow": add_cumulative_columns(market_flow),
        "stockFlows": {
            ticker: {
                metric: frame_to_records(frame)
                for metric, frame in frames.items()
            }
            for ticker, frames in stock_flows.items()
        },
    }


def latest_date_label(frame: pd.DataFrame) -> str:
    if frame.empty:
        return "-"
    return frame["日期"].max().strftime("%Y-%m-%d")


def frame_to_records(frame: pd.DataFrame) -> list[dict]:
    if frame.empty:
        return []
    normalized = frame.copy()
    for column in normalized.columns:
        if pd.api.types.is_datetime64_any_dtype(normalized[column]):
            normalized[column] = normalized[column].dt.strftime("%Y-%m-%d")
    return normalized.to_dict(orient="records")


def add_cumulative_columns(frame: pd.DataFrame) -> list[dict]:
    if frame.empty:
        return []
    view = frame.copy()
    for investor in ["外资", "机构", "散户"]:
        view[f"{investor}累计"] = view[investor].cumsum()
    return frame_to_records(view)


def build_html(data_json: str, meta: dict[str, str]) -> str:
    title = f"{meta['market']} 韩股资金流"
    escaped_title = html.escape(title)
    escaped_range = html.escape(f"{meta['start']} - {meta['end']}")
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escaped_title}</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f7f8f5;
      --panel: #ffffff;
      --ink: #17201c;
      --muted: #62706a;
      --line: #dfe6df;
      --foreign: #0f766e;
      --institution: #b45309;
      --retail: #2563eb;
      --soft: #eef4f0;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
    }}
    header {{
      padding: 28px clamp(16px, 4vw, 48px) 18px;
      border-bottom: 1px solid var(--line);
      background: #fbfcf9;
    }}
    h1 {{
      margin: 0;
      font-size: clamp(26px, 4vw, 42px);
      line-height: 1.12;
      letter-spacing: 0;
    }}
    .meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 14px;
      color: var(--muted);
      font-size: 14px;
    }}
    .meta span {{
      border: 1px solid var(--line);
      background: var(--panel);
      padding: 6px 10px;
      border-radius: 6px;
    }}
    main {{
      width: min(1220px, calc(100% - 32px));
      margin: 24px auto 42px;
      display: grid;
      gap: 22px;
    }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
    }}
    .section-head {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
      padding: 16px 18px;
      border-bottom: 1px solid var(--line);
      background: #fbfcf9;
    }}
    h2 {{
      margin: 0;
      font-size: 17px;
      line-height: 1.25;
      letter-spacing: 0;
    }}
    .hint {{ color: var(--muted); font-size: 13px; white-space: nowrap; }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
      padding: 18px;
    }}
    .kpi {{
      min-height: 112px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      background: var(--soft);
    }}
    .kpi .label {{ color: var(--muted); font-size: 13px; }}
    .kpi .value {{ margin-top: 8px; font-size: clamp(24px, 3vw, 34px); font-weight: 750; }}
    .kpi .state {{ margin-top: 8px; color: var(--muted); font-size: 13px; }}
    .chart-wrap {{ padding: 18px; }}
    canvas {{
      display: block;
      width: 100%;
      height: 420px;
    }}
    .tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      padding: 14px 18px 0;
    }}
    button {{
      min-height: 34px;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--ink);
      border-radius: 6px;
      padding: 7px 11px;
      cursor: pointer;
      font: inherit;
      font-size: 13px;
    }}
    button.active {{
      background: var(--ink);
      color: #fff;
      border-color: var(--ink);
    }}
    .table-wrap {{
      overflow: auto;
      padding: 18px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 740px;
      font-size: 13px;
    }}
    th, td {{
      padding: 10px 11px;
      border-bottom: 1px solid var(--line);
      text-align: right;
      white-space: nowrap;
    }}
    th:first-child, td:first-child,
    th:nth-child(2), td:nth-child(2) {{
      text-align: left;
    }}
    th {{
      background: #fbfcf9;
      color: #35433e;
      font-weight: 700;
      position: sticky;
      top: 0;
    }}
    .positive {{ color: #047857; }}
    .negative {{ color: #b91c1c; }}
    @media (max-width: 760px) {{
      .kpis {{ grid-template-columns: 1fr; }}
      canvas {{ height: 330px; }}
      .section-head {{ align-items: flex-start; flex-direction: column; }}
      .hint {{ white-space: normal; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>{escaped_title}</h1>
    <div class="meta">
      <span>区间 {escaped_range}</span>
      <span>最新交易日 {html.escape(meta["latestDate"])}</span>
      <span>生成时间 {html.escape(meta["generatedAt"])}</span>
    </div>
  </header>
  <main>
    <section>
      <div class="section-head">
        <h2>市场资金方向</h2>
        <div class="hint">单位：亿韩元，正数为净买入</div>
      </div>
      <div class="kpis" id="kpis"></div>
    </section>

    <section>
      <div class="section-head">
        <h2>累计资金流</h2>
        <div class="hint">外资、机构、散户累计净买卖金额</div>
      </div>
      <div class="chart-wrap"><canvas id="flowChart"></canvas></div>
    </section>

    <section>
      <div class="section-head">
        <h2>最近明细</h2>
        <div class="hint">市场与个股数据按日期倒序</div>
      </div>
      <div class="tabs" id="tabs"></div>
      <div class="table-wrap" id="tableWrap"></div>
    </section>
  </main>

  <script id="dashboard-data" type="application/json">{data_json}</script>
  <script>
    const data = JSON.parse(document.getElementById("dashboard-data").textContent);
    const investors = ["外资", "机构", "散户"];
    const colors = {{"外资": "#0f766e", "机构": "#b45309", "散户": "#2563eb"}};

    function formatNumber(value, digits = 0) {{
      return Number(value || 0).toLocaleString("zh-CN", {{
        maximumFractionDigits: digits,
        minimumFractionDigits: digits
      }});
    }}

    function className(value) {{
      const number = Number(value || 0);
      if (number > 0) return "positive";
      if (number < 0) return "negative";
      return "";
    }}

    function renderKpis() {{
      const host = document.getElementById("kpis");
      const rows = data.summary.filter(row => row["对象"] === "KOSPI市场");
      host.innerHTML = rows.map(row => `
        <div class="kpi">
          <div class="label">${{row["指标"]}}</div>
          <div class="value ${{className(row["最近值"])}}">${{formatNumber(row["最近值"], 2)}}</div>
          <div class="state">${{row["连续状态"]}}</div>
        </div>
      `).join("");
    }}

    function drawChart() {{
      const canvas = document.getElementById("flowChart");
      const ctx = canvas.getContext("2d");
      const ratio = window.devicePixelRatio || 1;
      const rect = canvas.getBoundingClientRect();
      canvas.width = rect.width * ratio;
      canvas.height = rect.height * ratio;
      ctx.setTransform(ratio, 0, 0, ratio, 0, 0);
      const width = rect.width;
      const height = rect.height;
      const pad = {{left: 58, right: 18, top: 24, bottom: 44}};
      const points = data.marketFlow;
      ctx.clearRect(0, 0, width, height);
      ctx.fillStyle = "#ffffff";
      ctx.fillRect(0, 0, width, height);
      if (!points.length) return;

      const values = points.flatMap(row => investors.map(name => Number(row[`${{name}}累计`] || 0)));
      let min = Math.min(...values, 0);
      let max = Math.max(...values, 0);
      if (min === max) {{ min -= 1; max += 1; }}
      const plotW = width - pad.left - pad.right;
      const plotH = height - pad.top - pad.bottom;
      const x = index => pad.left + (points.length === 1 ? plotW / 2 : (index / (points.length - 1)) * plotW);
      const y = value => pad.top + (max - value) / (max - min) * plotH;

      ctx.strokeStyle = "#dfe6df";
      ctx.lineWidth = 1;
      ctx.beginPath();
      for (let i = 0; i <= 4; i++) {{
        const gy = pad.top + i / 4 * plotH;
        ctx.moveTo(pad.left, gy);
        ctx.lineTo(width - pad.right, gy);
      }}
      ctx.stroke();

      ctx.fillStyle = "#62706a";
      ctx.font = "12px -apple-system, BlinkMacSystemFont, Segoe UI, sans-serif";
      ctx.textAlign = "right";
      ctx.textBaseline = "middle";
      for (let i = 0; i <= 4; i++) {{
        const value = max - i / 4 * (max - min);
        ctx.fillText(formatNumber(value, 0), pad.left - 8, pad.top + i / 4 * plotH);
      }}

      const zeroY = y(0);
      ctx.strokeStyle = "#94a39d";
      ctx.beginPath();
      ctx.moveTo(pad.left, zeroY);
      ctx.lineTo(width - pad.right, zeroY);
      ctx.stroke();

      investors.forEach(name => {{
        ctx.strokeStyle = colors[name];
        ctx.lineWidth = 2.5;
        ctx.beginPath();
        points.forEach((row, index) => {{
          const px = x(index);
          const py = y(Number(row[`${{name}}累计`] || 0));
          if (index === 0) ctx.moveTo(px, py);
          else ctx.lineTo(px, py);
        }});
        ctx.stroke();
      }});

      ctx.textAlign = "left";
      ctx.textBaseline = "alphabetic";
      investors.forEach((name, index) => {{
        const lx = pad.left + index * 92;
        const ly = height - 14;
        ctx.fillStyle = colors[name];
        ctx.fillRect(lx, ly - 10, 18, 3);
        ctx.fillStyle = "#35433e";
        ctx.fillText(name, lx + 25, ly - 5);
      }});

      ctx.fillStyle = "#62706a";
      ctx.textAlign = "left";
      ctx.fillText(points[0]["日期"], pad.left, height - 26);
      ctx.textAlign = "right";
      ctx.fillText(points[points.length - 1]["日期"], width - pad.right, height - 26);
    }}

    function tableFor(label, rows) {{
      const recent = [...rows].reverse().slice(0, 30);
      const columns = Object.keys(recent[0] || {{}});
      document.getElementById("tableWrap").innerHTML = `
        <table>
          <thead><tr>${{columns.map(col => `<th>${{col}}</th>`).join("")}}</tr></thead>
          <tbody>
            ${{recent.map(row => `<tr>${{columns.map(col => {{
              const value = row[col];
              const numeric = typeof value === "number";
              return `<td class="${{numeric ? className(value) : ""}}">${{numeric ? formatNumber(value, col.includes("累计") || label.includes("市场") ? 2 : 0) : value}}</td>`;
            }}).join("")}}</tr>`).join("")}}
          </tbody>
        </table>
      `;
    }}

    function renderTabs() {{
      const tabs = [
        {{label: "市场资金", rows: data.marketFlow}},
        ...Object.entries(data.stockFlows).flatMap(([ticker, frames]) => {{
          const stockTabs = [];
          if (frames.shares?.length) stockTabs.push({{label: `${{ticker}} 股数`, rows: frames.shares}});
          if (frames.value?.length) stockTabs.push({{label: `${{ticker}} 金额`, rows: frames.value}});
          return stockTabs;
        }})
      ];
      const host = document.getElementById("tabs");
      host.innerHTML = tabs.map((tab, index) => `<button type="button" data-index="${{index}}" class="${{index === 0 ? "active" : ""}}">${{tab.label}}</button>`).join("");
      host.addEventListener("click", event => {{
        const button = event.target.closest("button");
        if (!button) return;
        host.querySelectorAll("button").forEach(item => item.classList.remove("active"));
        button.classList.add("active");
        const tab = tabs[Number(button.dataset.index)];
        tableFor(tab.label, tab.rows);
      }});
      tableFor(tabs[0].label, tabs[0].rows);
    }}

    renderKpis();
    drawChart();
    renderTabs();
    window.addEventListener("resize", drawChart);
  </script>
</body>
</html>
"""


def format_sheet(sheet) -> None:
    if sheet.max_row < 1 or sheet.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        letter = column_cells[0].column_letter
        if header == "日期":
            sheet.column_dimensions[letter].width = 14
            for cell in column_cells[1:]:
                cell.number_format = "yyyy-mm-dd"
        elif header in {"对象", "指标", "连续状态"}:
            sheet.column_dimensions[letter].width = 20
        elif header == "单位":
            sheet.column_dimensions[letter].width = 10
        else:
            sheet.column_dimensions[letter].width = 14
            for cell in column_cells[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'


def resolve_tickers(client: KrxClient, requested: list[str]) -> dict[str, StockInfo]:
    stocks = {stock.ticker: stock for stock in client.listed_stocks()}
    resolved = {}
    missing = []
    for ticker in requested:
        if ticker in stocks:
            resolved[ticker] = stocks[ticker]
        else:
            missing.append(ticker)
    if missing:
        raise ValueError(f"KRX 没找到这些股票代码: {', '.join(missing)}")
    return resolved


def main() -> None:
    args = parse_args()
    tickers = [ticker.strip() for ticker in args.tickers.split(",") if ticker.strip()]
    output_path = Path(args.output or f"reports/krx_money_flow_{args.start}_{args.end}.xlsx")
    html_path = Path(args.html_output) if args.html_output else output_path.with_suffix(".html")
    chart_path = output_path.with_suffix(".market_flow.png")

    try:
        client = KrxClient(
            login_id=args.krx_id,
            login_password=args.krx_password,
            cookie_header=args.krx_cookie,
        )
        stock_info = resolve_tickers(client, tickers)

        market_flow = client.market_daily_flow(
            args.start,
            args.end,
            market=args.market,
            include_etf=args.include_etf,
            include_etn=args.include_etn,
            include_elw=args.include_elw,
        )

        stock_flows: dict[str, dict[str, pd.DataFrame]] = {}
        for ticker, info in stock_info.items():
            display_name = DEFAULT_TICKERS.get(ticker, info.name)
            key = f"{ticker}_{display_name}"
            stock_flows[key] = {
                "shares": client.stock_daily_flow(args.start, args.end, info.isin, metric="shares"),
                "value": client.stock_daily_flow(args.start, args.end, info.isin, metric="value"),
            }

        summary = build_summary(market_flow, stock_flows, market=args.market)
        payload = build_dashboard_payload(
            market_flow,
            stock_flows,
            summary,
            start=args.start,
            end=args.end,
            market=args.market,
        )
        plot_market_flow(market_flow, chart_path)
        write_report(output_path, market_flow, stock_flows, summary, chart_path)
        write_html_dashboard(
            html_path,
            market_flow,
            stock_flows,
            summary,
            start=args.start,
            end=args.end,
            market=args.market,
        )
        site_data_path = None
        if not args.no_site:
            site_data_path = write_site_data(Path(args.site_dir), payload)
    except Exception as exc:
        print(f"错误: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc

    print(f"已生成: {output_path}")
    print(f"已生成网页: {html_path}")
    if site_data_path is not None:
        print(f"已生成网站数据: {site_data_path}")


if __name__ == "__main__":
    main()
